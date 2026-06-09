import os
from collections import Counter, defaultdict
from typing import List, Dict, Any

import pandas as pd
from loguru import logger
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reddit_models import RedditPost, RedditComment
from config import EXPORTS_DIR, SMALL_DATASET_MIN_POSTS, SMALL_DATASET_MIN_COMMENTS, SMALL_DATASET_WARNING
from utils.date_utils import now_file_str

TOP_POSTS_COUNT = 30


# ─── Styling helpers ──────────────────────────────────────────────────────────

def _auto_fit(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        best = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                best = max(best, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(best + 4, max_width)


def _style_header(ws, hex_color: str = "2D5F8A") -> None:
    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=10)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28


def _style_section_row(ws, row_idx: int, hex_color: str = "D6E4F0") -> None:
    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    font = Font(bold=True, size=10)
    for cell in ws[row_idx]:
        if cell.value:
            cell.fill = fill
            cell.font = font


def _priority_color(priority: str) -> str:
    return {"high": "C6EFCE", "medium": "FFEB9C", "low": "FFFFFF"}.get(priority, "FFFFFF")


def _apply_priority_colors(ws, priority_col_idx: int) -> None:
    for row in ws.iter_rows(min_row=2):
        cell = row[priority_col_idx - 1]
        val = str(cell.value or "").lower()
        color = _priority_color(val)
        if color != "FFFFFF":
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for c in row:
                c.fill = fill


# ─── Quality Check ────────────────────────────────────────────────────────────

def _build_quality_rows(
    posts: List[RedditPost],
    comments: List[RedditComment],
    duplicate_posts_removed: int,
    all_subreddits: List[str] = None,
) -> List[dict]:
    def r(metric, value=""):
        return {"metric": metric, "value": value}

    total_p = len(posts)
    total_c = len(comments)
    empty_self = sum(1 for p in posts if not p.selftext.strip())
    bot_c = sum(1 for c in comments if c.is_bot_comment)
    low_score_p = sum(1 for p in posts if p.score < 5)
    no_kw_c = sum(1 for c in comments if c.comment_match_type == "no_match")
    avg_cmt = round(sum(p.num_comments for p in posts) / total_p, 1) if total_p else 0
    avg_ps = round(sum(p.score for p in posts) / total_p, 1) if total_p else 0
    avg_cs = round(sum(c.score for c in comments) / total_c, 1) if total_c else 0

    posts_per_sub = Counter(p.subreddit for p in posts)
    comments_per_sub = Counter(c.subreddit for c in comments)
    post_lang = Counter(p.language_detected for p in posts)
    comment_lang = Counter(c.language_detected for c in comments)

    pain_dist = Counter(p.pain_signal for p in posts)
    priority_dist = Counter(p.analysis_priority for p in posts)
    content_dist = Counter(p.content_type for p in posts)

    rows = [
        r("OVERVIEW"),
        r("total_posts", total_p),
        r("total_comments", total_c),
        r("duplicate_posts_removed", duplicate_posts_removed),
        r(""),
        r("DATA QUALITY"),
        r("empty_selftext_count", empty_self),
        r("bot_comments_filtered", bot_c),
        r("low_score_posts (score < 5)", low_score_p),
        r("comments_no_keyword_match", no_kw_c),
        r("average_comments_per_post", avg_cmt),
        r("average_post_score", avg_ps),
        r("average_comment_score", avg_cs),
    ]

    if total_p < SMALL_DATASET_MIN_POSTS or total_c < SMALL_DATASET_MIN_COMMENTS:
        rows.append(r("DATASET WARNING", SMALL_DATASET_WARNING))

    rows += [r(""), r("ANALYSIS PRIORITY DISTRIBUTION")]
    for pri in ("high", "medium", "low"):
        rows.append(r(f"  {pri}", priority_dist.get(pri, 0)))

    rows += [r(""), r("PAIN SIGNAL DISTRIBUTION")]
    for signal, cnt in pain_dist.most_common():
        rows.append(r(f"  {signal}", cnt))

    rows += [r(""), r("CONTENT TYPE DISTRIBUTION")]
    for ctype, cnt in content_dist.most_common():
        rows.append(r(f"  {ctype}", cnt))

    rows += [r(""), r("POSTS PER SUBREDDIT")]
    for sub, cnt in sorted(posts_per_sub.items(), key=lambda x: -x[1]):
        rows.append(r(f"  r/{sub}", cnt))

    rows += [r(""), r("COMMENTS PER SUBREDDIT")]
    for sub, cnt in sorted(comments_per_sub.items(), key=lambda x: -x[1]):
        rows.append(r(f"  r/{sub}", cnt))

    rows += [r(""), r("POST LANGUAGE DISTRIBUTION")]
    for lang, cnt in sorted(post_lang.items()):
        rows.append(r(f"  {lang}", cnt))

    rows += [r(""), r("COMMENT LANGUAGE DISTRIBUTION")]
    for lang, cnt in sorted(comment_lang.items()):
        rows.append(r(f"  {lang}", cnt))

    # Subreddits with zero posts collected
    if all_subreddits:
        subs_with_posts = set(p.subreddit for p in posts)
        zero_subs = [s for s in all_subreddits if s not in subs_with_posts]
        rows += [r(""), r("SUBREDDITS WITH ZERO RESULTS")]
        if zero_subs:
            for s in zero_subs:
                rows.append(r(f"  r/{s}", "0 posts"))
        else:
            rows.append(r("  (all subreddits returned results)", ""))

    return rows


# ─── Top Posts sheet ──────────────────────────────────────────────────────────

def _build_top_posts(posts: List[RedditPost], n: int = TOP_POSTS_COUNT) -> List[dict]:
    sorted_posts = sorted(posts, key=lambda p: p.trend_score, reverse=True)[:n]
    rows = []
    for rank, p in enumerate(sorted_posts, 1):
        short = (p.selftext[:200] + "…") if len(p.selftext) > 200 else p.selftext
        rows.append({
            "rank": rank,
            "subreddit": f"r/{p.subreddit}",
            "title": p.title,
            "score": p.score,
            "num_comments": p.num_comments,
            "trend_score": p.trend_score,
            "analysis_priority": p.analysis_priority,
            "pain_signal": p.pain_signal,
            "content_type": p.content_type,
            "matched_keywords": p.matched_keywords,
            "permalink": p.permalink,
            "short_selftext": short,
        })
    return rows


# ─── Keyword Summary sheet ────────────────────────────────────────────────────

def _build_keyword_summary(
    posts: List[RedditPost],
    comments: List[RedditComment],
) -> List[dict]:
    post_kw: Counter = Counter()
    kw_post_subs: Dict[str, Counter] = defaultdict(Counter)
    kw_comment_subs: Dict[str, Counter] = defaultdict(Counter)

    for p in posts:
        for kw in p.matched_keywords.split(", "):
            kw = kw.strip()
            if kw:
                post_kw[kw] += 1
                kw_post_subs[kw][p.subreddit] += 1

    comment_kw: Counter = Counter()
    for c in comments:
        for kw in c.matched_keywords.split(", "):
            kw = kw.strip()
            if kw:
                comment_kw[kw] += 1
                kw_comment_subs[kw][c.subreddit] += 1

    all_kw = set(post_kw) | set(comment_kw)
    rows = []
    for kw in sorted(all_kw, key=lambda k: -(post_kw.get(k, 0) + comment_kw.get(k, 0))):
        pc = post_kw.get(kw, 0)
        cc = comment_kw.get(kw, 0)
        # Use post subreddits first; fall back to comment subreddits if keyword only in comments
        sub_source = kw_post_subs[kw] if kw_post_subs[kw] else kw_comment_subs[kw]
        top_subs = ", ".join(f"r/{s}({n})" for s, n in sub_source.most_common(3))
        rows.append({
            "keyword": kw,
            "posts_count": pc,
            "comments_count": cc,
            "total_mentions": pc + cc,
            "top_subreddits": top_subs,
        })
    return rows


# ─── Main export ──────────────────────────────────────────────────────────────

def export_excel(
    posts: List[RedditPost],
    comments: List[RedditComment],
    run_settings: Dict[str, Any],
    output_path: str = None,
    duplicate_posts_removed: int = 0,
    all_subreddits: List[str] = None,
) -> str:
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    if not output_path:
        output_path = os.path.join(EXPORTS_DIR, f"reddit_{now_file_str()}.xlsx")

    small = len(posts) < SMALL_DATASET_MIN_POSTS or len(comments) < SMALL_DATASET_MIN_COMMENTS

    summary = {
        "run_date": run_settings.get("run_date", ""),
        "subreddits": run_settings.get("subreddits", ""),
        "keywords": run_settings.get("keywords", ""),
        "period": run_settings.get("period", ""),
        "sort": run_settings.get("sort", ""),
        "total_posts": len(posts),
        "total_comments": len(comments),
        "min_score": run_settings.get("min_score", 0),
        "min_comments": run_settings.get("min_comments", 0),
        "max_comments_per_post": run_settings.get("max_comments_per_post", 0),
        "language_mode": run_settings.get("language_mode", "mixed"),
        "dataset_status": ("WARNING — " + SMALL_DATASET_WARNING) if small else "OK",
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # 1. Summary
        pd.DataFrame([summary]).to_excel(writer, sheet_name="Summary", index=False)
        ws = writer.sheets["Summary"]
        _style_header(ws, "2D5F8A")
        _auto_fit(ws)

        # 2. Top Posts
        top_rows = _build_top_posts(posts)
        df_top = pd.DataFrame(top_rows) if top_rows else pd.DataFrame(columns=[
            "rank", "subreddit", "title", "score", "num_comments", "trend_score",
            "analysis_priority", "pain_signal", "content_type",
            "matched_keywords", "permalink", "short_selftext",
        ])
        df_top.to_excel(writer, sheet_name="Top Posts", index=False)
        ws = writer.sheets["Top Posts"]
        _style_header(ws, "1A5276")
        # Color rows by analysis_priority (col index 7 = analysis_priority)
        priority_col = df_top.columns.get_loc("analysis_priority") + 1
        _apply_priority_colors(ws, priority_col)
        _auto_fit(ws)

        # 3. Posts
        posts_data = [p.to_dict() for p in posts]
        df_posts = pd.DataFrame(posts_data) if posts_data else pd.DataFrame(
            columns=list(RedditPost.__dataclass_fields__.keys()))
        df_posts.to_excel(writer, sheet_name="Posts", index=False)
        ws = writer.sheets["Posts"]
        _style_header(ws, "2D5F8A")
        _auto_fit(ws)

        # 4. Comments
        comments_data = [c.to_dict() for c in comments]
        df_comments = pd.DataFrame(comments_data) if comments_data else pd.DataFrame(
            columns=list(RedditComment.__dataclass_fields__.keys()))
        df_comments.to_excel(writer, sheet_name="Comments", index=False)
        ws = writer.sheets["Comments"]
        _style_header(ws, "2D5F8A")
        _auto_fit(ws)

        # 5. Keyword Summary
        kw_rows = _build_keyword_summary(posts, comments)
        df_kw = pd.DataFrame(kw_rows) if kw_rows else pd.DataFrame(columns=[
            "keyword", "posts_count", "comments_count", "total_mentions", "top_subreddits"])
        df_kw.to_excel(writer, sheet_name="Keyword Summary", index=False)
        ws = writer.sheets["Keyword Summary"]
        _style_header(ws, "117A65")
        _auto_fit(ws)

        # 6. Quality Check
        q_rows = _build_quality_rows(posts, comments, duplicate_posts_removed, all_subreddits)
        df_q = pd.DataFrame(q_rows)
        df_q.to_excel(writer, sheet_name="Quality Check", index=False)
        ws = writer.sheets["Quality Check"]
        _style_header(ws, "4A7C59")
        # Style section header rows (non-empty metric with empty value)
        for row in ws.iter_rows(min_row=2):
            metric_cell = row[0]
            value_cell = row[1] if len(row) > 1 else None
            val = str(metric_cell.value or "").strip()
            is_section = val and (not value_cell or not str(value_cell.value or "").strip())
            if is_section and val != "":
                _style_section_row(ws, metric_cell.row)
        _auto_fit(ws)

        # 7. Run Settings
        settings_rows = [{"parameter": k, "value": str(v)} for k, v in run_settings.items()]
        pd.DataFrame(settings_rows).to_excel(writer, sheet_name="Run Settings", index=False)
        ws = writer.sheets["Run Settings"]
        _style_header(ws, "2D5F8A")
        _auto_fit(ws)

    if small:
        logger.warning(SMALL_DATASET_WARNING)

    logger.success(f"Excel exported: {output_path}")
    return output_path
